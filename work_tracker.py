import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import json
from pathlib import Path

class WorkTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Personal Work Tracker")
        self.root.geometry("900x500")
        
        # Set up data directory
        self.data_dir = Path.home() / "WorkTrackerData"
        self.data_dir.mkdir(exist_ok=True)
        self.tasks_file = self.data_dir / "tasks.json"
        
        # Project categories and their keywords
        self.project_categories = {
            "GSU meeting/Training": ["gsu", "meeting", "training"],
            "Onboarding": ["onboard", "orientation"],
            "Life Review": ["life", "review"],
            "LASSI Review": ["lassi"],
            "Recruitment": ["recruit", "admission"],
            "Student": ["student"],
            "Coaching": ["coach"],
            "Tutoring": ["tutor"],
            "Trio Project Assistance": ["trio", "project", "assist"],
            "Administrative": ["admin", "paperwork", "documentation"],
            "Trio Team Meeting": ["team", "meeting"],
            "Work Assistance": ["work", "assist"],
            "Group Coaching": ["group", "coach"],
            "Cultural Events": ["cultural", "event"],
            "Team Collaboration": ["collaboration", "collab"],
            "Success Workshop": ["success", "workshop"],
            "Trio Training": ["trio", "training"],
            "Other": []
        }
        
        # List to store all tasks
        self.tasks = []
        
        # Current task tracking
        self.current_task = None
        self.current_start_time = None
        self.current_date = None
        self.current_project = None
        
        # Set up the GUI first
        self.setup_gui()
        
        # Then load tasks and populate the treeview
        self.load_tasks()

    def setup_gui(self):
        # Create main frame with padding
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Current Task Status Frame
        self.status_frame = ttk.LabelFrame(self.main_frame, text="Current Task Status", padding="5")
        self.status_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        self.current_task_label = ttk.Label(self.status_frame, text="No task in progress")
        self.current_task_label.pack(fill=tk.X)
        
        # Task Entry
        ttk.Label(self.main_frame, text="Current Task:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.task_entry = ttk.Entry(self.main_frame, width=50)
        self.task_entry.grid(row=1, column=1, columnspan=2, sticky=tk.W, pady=5)
        self.task_entry.bind('<KeyRelease>', self.suggest_project)
        
        # Project ID Dropdown
        ttk.Label(self.main_frame, text="Project ID:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.project_id = ttk.Combobox(self.main_frame, width=30, values=list(self.project_categories.keys()))
        self.project_id.grid(row=2, column=1, sticky=tk.W, pady=5)
        
        # Time Entry Frames
        time_frame = ttk.Frame(self.main_frame)
        time_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # Start Time
        start_frame = ttk.LabelFrame(time_frame, text="Start Time")
        start_frame.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        self.start_time_entry = ttk.Entry(start_frame, width=20)
        self.start_time_entry.pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Label(start_frame, text="(HH:MM AM/PM)").pack(side=tk.LEFT, padx=2)
        ttk.Button(start_frame, text="Use Current", 
                  command=lambda: self.set_current_time(self.start_time_entry)).pack(side=tk.LEFT, padx=5, pady=5)
        
        # End Time
        end_frame = ttk.LabelFrame(time_frame, text="End Time")
        end_frame.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        self.end_time_entry = ttk.Entry(end_frame, width=20)
        self.end_time_entry.pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Label(end_frame, text="(HH:MM AM/PM)").pack(side=tk.LEFT, padx=2)
        ttk.Button(end_frame, text="Use Current", 
                  command=lambda: self.set_current_time(self.end_time_entry)).pack(side=tk.LEFT, padx=5, pady=5)
        
        # Create frame for Treeview
        tree_frame = ttk.Frame(self.main_frame)
        tree_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # Treeview for tasks
        self.tree = ttk.Treeview(tree_frame, 
                                columns=('Date', 'Project ID', 'Task', 'Start Time', 'End Time', 'Duration'),
                                show='headings', 
                                height=10)
        
        # Add right-click menu
        self.context_menu = tk.Menu(self.tree, tearoff=0)
        self.context_menu.add_command(label="Edit Task", command=self.edit_selected_task)
        self.context_menu.add_command(label="Delete Task", command=self.delete_selected_task)
        self.tree.bind("<Button-3>", self.show_context_menu)  # Right-click
        self.tree.bind("<Delete>", lambda e: self.delete_selected_task())  # Delete key
        self.tree.bind("<Double-1>", lambda e: self.edit_selected_task())  # Double-click to edit
        
        # Define headings
        self.tree.heading('Date', text='Date')
        self.tree.heading('Project ID', text='Project ID')
        self.tree.heading('Task', text='Task')
        self.tree.heading('Start Time', text='Start Time')
        self.tree.heading('End Time', text='End Time')
        self.tree.heading('Duration', text='Duration')
        
        # Define columns
        self.tree.column('Date', width=100)
        self.tree.column('Project ID', width=150)
        self.tree.column('Task', width=250)
        self.tree.column('Start Time', width=120)
        self.tree.column('End Time', width=120)
        self.tree.column('Duration', width=120)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Grid the treeview and scrollbar
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Buttons at the bottom
        button_frame = ttk.Frame(self.main_frame)
        button_frame.grid(row=5, column=0, columnspan=3, pady=10)
        
        ttk.Button(button_frame, text="Start Task", 
                  command=self.start_task).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Complete Task", 
                  command=self.complete_task).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Generate Report", 
                  command=self.generate_report).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear All Tasks", 
                  command=self.clear_all_tasks).pack(side=tk.LEFT, padx=5)

    def load_tasks(self):
        """Load tasks and current task status from JSON file"""
        try:
            if self.tasks_file.exists():
                with open(self.tasks_file, 'r') as f:
                    data = json.load(f)
                    self.tasks = data.get('completed_tasks', [])
                    current_task = data.get('current_task', {})
                    if current_task:
                        self.current_task = current_task.get('task')
                        self.current_start_time = current_task.get('start_time')
                        self.current_date = current_task.get('date')
                        self.current_project = current_task.get('project')
                        if self.current_project:
                            self.project_id.set(self.current_project)
                        self.update_current_task_display()
            # After loading, populate the treeview
            self.populate_treeview()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load tasks: {str(e)}")
            self.tasks = []

    def save_tasks(self):
        """Save tasks and current task status to JSON file"""
        try:
            data = {
                'completed_tasks': self.tasks,
                'current_task': {
                    'task': self.current_task,
                    'start_time': self.current_start_time,
                    'date': self.current_date,
                    'project': self.current_project
                } if self.current_task else {}
            }
            with open(self.tasks_file, 'w') as f:
                json.dump(data, f, indent=4)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save tasks: {str(e)}")

    def show_context_menu(self, event):
        """Show right-click menu"""
        try:
            # Select row under mouse
            item = self.tree.identify_row(event.y)
            if item:
                self.tree.selection_set(item)
                self.context_menu.post(event.x_root, event.y_root)
        finally:
            # Required to make menu disappear when clicking outside
            self.context_menu.grab_release()

    def delete_selected_task(self):
        """Delete the selected task from both treeview and tasks list"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select a task to delete")
            return

        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this task?"):
            for item in selected_items:
                # Get values from the selected row
                values = self.tree.item(item)['values']
                
                # Find and remove the task from self.tasks
                for task in self.tasks[:]:  # Create a copy of the list to iterate over
                    if (task['date'] == values[0] and 
                        task['project_id'] == values[1] and 
                        task['task'] == values[2] and 
                        task['start_time'] == values[3] and 
                        task['end_time'] == values[4]):
                        self.tasks.remove(task)
                        break
                
                # Remove from treeview
                self.tree.delete(item)
            
            # Save updated tasks
            self.save_tasks()
            messagebox.showinfo("Success", "Task deleted successfully")

    def populate_treeview(self):
        """Populate treeview with loaded tasks"""
        try:
            # Clear existing items
            for item in self.tree.get_children():
                self.tree.delete(item)
                
            # Add tasks to treeview
            for task in self.tasks:
                self.tree.insert('', tk.END, values=(
                    task['date'],
                    task['project_id'],
                    task['task'],
                    task['start_time'],
                    task['end_time'],
                    task['duration']
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to populate task list: {str(e)}")

    def set_current_time(self, entry_widget):
        """Set the current time in the entry widget"""
        current_time = datetime.now().strftime("%I:%M %p")  # Changed to HH:MM AM/PM format
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, current_time)

    def validate_time(self, time_str):
        """Validate time string in either HH:MM AM/PM or HH:MM:SS AM/PM format"""
        time_formats = [
            "%I:%M %p",     # 12-hour without seconds (e.g., "10:20 AM")
            "%H:%M",         # 24-hour without seconds
            "%I:%M:%S %p",  # 12-hour with seconds (e.g., "10:20:11 AM")
            "%H:%M:%S"      # 24-hour with seconds
        ]
        
        for time_format in time_formats:
            try:
                return datetime.strptime(time_str, time_format)
            except ValueError:
                continue
        
        raise ValueError("Invalid time format! Please use HH:MM AM/PM or HH:MM:SS AM/PM")

    def suggest_project(self, event=None):
        task_text = self.task_entry.get().lower()
        
        # Check each project category for matching keywords
        for project, keywords in self.project_categories.items():
            if any(keyword in task_text for keyword in keywords):
                self.project_id.set(project)
                break
        else:
            # If no matches found, set to "Other"
            self.project_id.set("Other")

    def update_current_task_display(self):
        """Update the current task status display"""
        if self.current_task:
            status_text = (f"In Progress: {self.current_task}\n"
                         f"Project: {self.current_project}\n"
                         f"Started at: {self.current_start_time} on {self.current_date}")
            self.current_task_label.config(
                text=status_text,
                foreground="green",
                font=("Arial", 10, "bold"),
                justify=tk.LEFT
            )
        else:
            self.current_task_label.config(
                text="No task in progress",
                foreground="black",
                font=("Arial", 10),
                justify=tk.LEFT
            )

    def start_task(self):
        # Check if there's already a task in progress
        if self.current_task:
            messagebox.showerror("Error", f"A task is already in progress: {self.current_task}\nPlease complete it first!")
            return
            
        task = self.task_entry.get().strip()
        start_time = self.start_time_entry.get().strip()
        project = self.project_id.get()
        
        if not task:
            messagebox.showerror("Error", "Please enter a task description!")
            return
            
        if not start_time:
            messagebox.showerror("Error", "Please enter start time!")
            return
            
        if not project:
            messagebox.showerror("Error", "Please select a Project ID!")
            return
            
        try:
            self.validate_time(start_time)
        except ValueError as e:
            messagebox.showerror("Error", str(e))
            return
            
        self.current_task = task
        self.current_start_time = start_time
        self.current_date = datetime.now().strftime("%m/%d/%Y")
        self.current_project = project  # Save the current project
        
        # Save the current task status
        self.save_tasks()
        
        # Update status display
        self.update_current_task_display()
        
        messagebox.showinfo("Success", f"Task '{task}' started at {start_time}")

    def complete_task(self):
        if not self.current_task:
            messagebox.showerror("Error", "No task is currently running!")
            return
        
        # Get end time from entry if provided, otherwise use current time
        end_time = self.end_time_entry.get().strip()
        if not end_time:
            end_time = datetime.now().strftime("%I:%M %p")
        else:
            try:
                self.validate_time(end_time)
            except ValueError as e:
                messagebox.showerror("Error", str(e))
                return
        
        # Calculate duration
        try:
            start_time_obj = self.validate_time(self.current_start_time)
            end_time_obj = self.validate_time(end_time)
            duration = end_time_obj - start_time_obj
            duration_str = f"{duration.seconds // 3600:02d}:{(duration.seconds % 3600) // 60:02d}"
        except ValueError as e:
            messagebox.showerror("Error", f"Error calculating duration: {str(e)}")
            return
        
        # Add to tasks list
        task_info = {
            'date': self.current_date,
            'project_id': self.current_project,  # Use saved project
            'task': self.current_task,
            'task_id': '',  # Empty
            'category': 'Administrative Task' if 'admin' in self.current_task.lower() else 'Other',
            'start_time': self.current_start_time,
            'end_time': end_time,
            'duration': duration_str,
            'breaks': '',  # Empty
            'minutes_worked': '',  # Empty
            'faculty_student_staff': ''  # Empty
        }
        self.tasks.append(task_info)
        
        # Save tasks to file
        self.save_tasks()
        
        # Add to treeview
        self.tree.insert('', tk.END, values=(
            self.current_date,
            self.current_project,  # Use saved project
            self.current_task,
            self.current_start_time,
            end_time,
            duration_str
        ))
        
        # Reset current task and clear entries
        self.current_task = None
        self.current_start_time = None
        self.current_date = None
        self.current_project = None  # Clear current project
        self.task_entry.delete(0, tk.END)
        self.start_time_entry.delete(0, tk.END)
        self.end_time_entry.delete(0, tk.END)
        self.project_id.set('')
        
        # Update status display
        self.update_current_task_display()
        
        messagebox.showinfo("Success", "Task completed!")

    def clear_all_tasks(self):
        """Clear all tasks from the treeview and storage"""
        if not self.tasks:
            messagebox.showinfo("Info", "No tasks to clear!")
            return
            
        if messagebox.askyesno("Confirm Clear All", 
                              "Are you sure you want to clear all tasks? This action cannot be undone!",
                              icon='warning'):
            # Clear treeview
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Clear tasks list
            self.tasks = []
            
            # Save empty task list
            self.save_tasks()
            
            # Reset current task if any
            self.current_task = None
            self.current_start_time = None
            self.current_date = None
            self.current_project = None  # Clear current project
            
            # Clear entry fields
            self.task_entry.delete(0, tk.END)
            self.start_time_entry.delete(0, tk.END)
            self.end_time_entry.delete(0, tk.END)
            self.project_id.set('')
            
            # Update status display
            self.update_current_task_display()
            
            messagebox.showinfo("Success", "All tasks have been cleared!")

    def edit_selected_task(self):
        """Open edit dialog for selected task"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select a task to edit")
            return

        # Get the selected task
        item = selected_items[0]
        task_values = self.tree.item(item)['values']

        # Create edit dialog
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Task")
        edit_window.geometry("500x300")

        # Add fields
        ttk.Label(edit_window, text="Task:").grid(row=0, column=0, padx=5, pady=5)
        task_entry = ttk.Entry(edit_window, width=40)
        task_entry.insert(0, task_values[2])
        task_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(edit_window, text="Project ID:").grid(row=1, column=0, padx=5, pady=5)
        project_combo = ttk.Combobox(edit_window, width=30, values=list(self.project_categories.keys()))
        project_combo.set(task_values[1])
        project_combo.grid(row=1, column=1, padx=5, pady=5)

        # Time Entry Frames
        time_frame = ttk.Frame(edit_window)
        time_frame.grid(row=2, column=0, columnspan=2, pady=10)

        ttk.Label(time_frame, text="Start Time:").pack(side=tk.LEFT, padx=5)
        start_time = ttk.Entry(time_frame, width=20)
        start_time.pack(side=tk.LEFT, padx=5)
        start_time.insert(0, task_values[3])

        ttk.Label(time_frame, text="End Time:").pack(side=tk.LEFT, padx=5)
        end_time = ttk.Entry(time_frame, width=20)
        end_time.pack(side=tk.LEFT, padx=5)
        end_time.insert(0, task_values[4])

        def save_changes():
            # Validate times
            try:
                self.validate_time(start_time.get())
                self.validate_time(end_time.get())
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid time format: {str(e)}")
                return

            # Calculate new duration
            try:
                start = self.validate_time(start_time.get())
                end = self.validate_time(end_time.get())
                duration = end - start
                duration_str = f"{duration.seconds // 3600:02d}:{(duration.seconds % 3600) // 60:02d}"
            except ValueError as e:
                messagebox.showerror("Error", f"Error calculating duration: {str(e)}")
                return

            # Update task in self.tasks list
            for task in self.tasks:
                if (task['date'] == task_values[0] and 
                    task['project_id'] == task_values[1] and 
                    task['task'] == task_values[2]):
                    
                    task['task'] = task_entry.get()
                    task['project_id'] = project_combo.get()
                    task['start_time'] = start_time.get()
                    task['end_time'] = end_time.get()
                    task['duration'] = duration_str
                    break

            # Update treeview
            self.tree.item(item, values=(
                task_values[0],  # Keep original date
                project_combo.get(),
                task_entry.get(),
                start_time.get(),
                end_time.get(),
                duration_str
            ))

            # Save changes to file
            self.save_tasks()
            
            edit_window.destroy()
            messagebox.showinfo("Success", "Task updated successfully")

        # Add Save button
        ttk.Button(edit_window, text="Save Changes", command=save_changes).grid(row=3, column=0, columnspan=2, pady=20)

    def generate_report(self):
        if not self.tasks:
            messagebox.showerror("Error", "No tasks to generate report!")
            return
            
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        headers = [
            'Date', 
            'Project ID', 
            'Task ID', 
            'Faculty Student or Staff', 
            'Administrative Task or Other', 
            'Start Time', 
            'Breaks (minutes)', 
            'End Time', 
            'Minutes Worked'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Add data
        for row, task in enumerate(self.tasks, 2):
            ws.cell(row=row, column=1).value = task['date']
            ws.cell(row=row, column=2).value = task['project_id']
            ws.cell(row=row, column=3).value = task['task_id']  # Empty
            ws.cell(row=row, column=4).value = task['faculty_student_staff']  # Empty
            ws.cell(row=row, column=5).value = task['category']
            ws.cell(row=row, column=6).value = task['start_time']
            ws.cell(row=row, column=7).value = task['breaks']  # Empty
            ws.cell(row=row, column=8).value = task['end_time']
            ws.cell(row=row, column=9).value = task['minutes_worked']  # Empty
            
            # Center align all columns
            for col in range(1, 10):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        
        # Set fixed column widths
        column_widths = {
            'A': 12,  # Date
            'B': 20,  # Project ID
            'C': 15,  # Task ID
            'D': 25,  # Faculty Student or Staff
            'E': 25,  # Administrative Task or Other
            'F': 15,  # Start Time
            'G': 15,  # Breaks
            'H': 15,  # End Time
            'I': 15   # Minutes Worked
        }
        
        # Apply fixed widths
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        filename = f"Rohit_Work_Summary_{datetime.now().strftime('%m_%d_%Y')}.xlsx"
        wb.save(filename)
        messagebox.showinfo("Success", f"Report generated successfully!\nSaved as: {filename}")

def main():
    root = tk.Tk()
    app = WorkTracker(root)
    root.mainloop()

if __name__ == "__main__":
    main()